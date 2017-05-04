# encoding:utf-8
import openpyxl
from order import OrderService
import os
from global_param import *
import datetime


class Exporter:
    def __init__(self):
        self.odsvs = OrderService()

    def _is_defective(self, defective):
        if defective == '0':
            return '是'
        else:
            return '-'

    def export(self,trans_day):
        wb = openpyxl.Workbook()

        total_amount = 0
        total_packages = 0




        model_price_sheet = wb.create_sheet('统计-型号-价格',0)
        model_price = self.odsvs.stat_by_model_price()
        model_price_sheet.append(['型号', '件数', '双数', '单价', '折扣', '总价'])
        for d in model_price:
            model_price_sheet.append([
                d['model'],
                d['packages'],
                d['pairs'],
                d['price'],
                d['discount'],
                d['amount']
            ])






        detail_sheet = wb.create_sheet('明细',1)
        detail = self.odsvs.get_all_orders()
        detail_sheet.append(['型号','件数','双数','单价','折扣','总价','次品'])
        for d in detail:
            detail_sheet.append([
                d['model'],
                d['packages'],
                d['pairs'],
                d['price'],
                d['discount'],
                d['amount'],
                self._is_defective(d['defective'])

            ])

        filename = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        wb.save(filename=EXPORT_FOLDER+'%s.xlsx' % filename)

        return '%s.xlsx' % filename




if __name__ == '__main__':
    e = Exporter()
    e.export('2017-04-21')