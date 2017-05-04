# encoding:utf-8
from openpyxl import load_workbook
from mysqlconn import MySQLHelper
import pandas as pd
import fire
from global_param import *
import datetime
# ./
# 2017年4月23日叶卡销售明细表.xlsx
# 2017年4月26日叶卡销售明细表.xlsx


class ExcelParser:
    def __init__(self,filename):
        self.filename = filename
        if not filename.endswith('.xls'):
            if not filename.endswith('.xlsx'):
                raise Exception('ONLY EXCEL FILE ACCEPT!')
        self.wb = load_workbook(filename=UPLOAD_FOLDER+ filename)

        self.mysql_helper = MySQLHelper()
        self.trans_day = filename.split('.')[0]




    def parser(self):
        sheetname = self.wb.sheetnames[0]
        sheet = self.wb.get_sheet_by_name(sheetname)

        data = list()
        for row in list(sheet.rows)[2:]:
            model,packages,pairs,price,discount = [c.value for c in list(row)[:5]]


            if model is None:
                break

            if '处理' in str(model):
                model = model.replace('处理','')
                defective = '0'
            else:
                defective = '1'


            data.append({'trans_day':self.trans_day,
                         'model':model,
                         'packages':packages,
                         'pairs':pairs,
                         'price':price,
                         'discount':discount,
                         'defective':defective})

        #pds = pd.DataFrame(data,['day','model','packages','pairs','price','discount'])

        pds = pd.DataFrame(data).fillna(value=0)




        pds['amount'] = pds.price * pds.pairs * pds.packages - pds.discount

        data = list()
        for i,r in pds.iterrows():
            data.append(['',r.model,r.packages,r.pairs,r.price,r.discount,r.amount,r.defective])


        # for d in data:
        #     print(d)


        self.mysql_helper.execute("truncate table t_order")

        sql = "insert into t_order(trans_day,model,packages,pairs,price,discount,amount,defective) values(%s,%s,%s,%s,%s,%s,%s,%s)"
        self.mysql_helper.executemany(sql,data)






if __name__ == '__main__':
    #fire.Fire(ExcelParser)
    parser = ExcelParser('2017年4月23日叶卡销售明细表.xlsx')
    parser.parser()